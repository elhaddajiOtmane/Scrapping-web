from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl as xl
import os
from selenium.webdriver.chrome.service import Service
os.chdir("D:\\OLD DATA\\finalcc")
wb = xl.load_workbook("D:\\OLD DATA\\finalcc\\codelist.xlsx")
ws = wb.active
service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)
time.sleep(2)
list_of_pages=[]
for i in range(2,6):
    column =i-1
    row =1
    print("Page no. = "+str(i))
    driver.get('https://codelist.cc/pgs/'+str(i)+'/')
    time.sleep(1)
    elems = driver.find_elements(By.XPATH,"//a[@href]")
    list_of_urls = []
    for elem in elems:
        check = str(elem.get_attribute("href"))
        print(check)
        if "#comment" in check:
            check = check.replace("#comment","")
        if((len(check) > 60) and (list_of_urls.count(check)==0)):
            print(check)
            if 'scripts' in check:
                list_of_urls.append(check)
                ws.cell(row=row,column=column).value = check
                row=row+1
            if 'plugins' in check:
                list_of_urls.append(check)
                ws.cell(row=row,column=column).value = check
                row=row+1
            if 'mobile' in check:
                list_of_urls.append(check)
                ws.cell(row=row,column=column).value = check
                row=row+1
    print(list_of_urls)
    list_of_pages.append(list_of_urls)
    wb.save("codelist.xlsx")
# print("List of all the pages",list_of_pages)
#     time.sleep(2)
#     list_of_pages.append(list_of_urls)
#     wb.save("codelist.xlsx")
print(list_of_pages)
# for elemt in elems:
#   x=elemt.get_attribute("href")
#   print("inside first site it is"+str(x))
#   driver.get(x)
#   time.sleep(2)
#   elemn = driver.find_elements(By.XPATH,"//a[@href]")
#   for ele in elemn:
#       print("printing urls")
#       print(ele.get_attribute("href"))
# try:
#   elemt='https://codelist.cc/disclaimer.html'
#   x=elemt.get_attribute("href")
# except Exception:
#   print("please move on")