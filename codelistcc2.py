from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl as xl
import os
def get_no_of_rows(worksheet):
    count=0
    for k in range(1,(worksheet.max_row)+1):
        if(worksheet.cell(row=k,column=1).value == None):
            count=count+1
    no_of_rows = (worksheet.max_row) - count
    return no_of_rows
def get_no_of_columns(worksheet):
    countc=0
    for l in range(1,(worksheet.max_column)+1):
        if(worksheet.cell(row=1,column=l).value == None):
            countc=countc+1
    no_of_cols = (worksheet.max_column) - countc
    return no_of_cols
from selenium.webdriver.chrome.service import Service
service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)
time.sleep(2)
os.chdir("D:\\OLD DATA\\finalcc")
wb = xl.load_workbook("D:\\OLD DATA\\finalcc\\codelist.xlsx")
ws = wb.active
rows = get_no_of_rows(ws)
columns = get_no_of_columns(ws)
for i in range(1,rows+1):
    print(i)
    for j in range(1,columns+1):
        link = ws.cell(row=i,column=j).value
        driver.get(link)
        time.sleep(2)
        print(j)
        elem=driver.find_element(By.CLASS_NAME,("quote"))
        innertext = elem.get_attribute('innerHTML')
        print("check")
        text = str(innertext.strip())
        ftext =text.replace("<!--QuoteEBegin-->"," ")
        ftext =ftext.replace("<!--QuoteEnd-->"," ")
        ftext =ftext.replace("<br>","\n")
        ws.cell(row=i,column=j).value = ftext
        wb.save("codelist2.xlsx")


      